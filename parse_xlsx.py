from openpyxl import load_workbook

wb = load_workbook(filename=r"Тест файл.xlsx")
ws = wb.active
m_row = ws.max_row
m_col = ws.max_column
for i in range(1, m_col + 1):
    for n in range(1, m_row + 1):
        cell_obj = ws.cell(row=n, column=i)
        print(cell_obj.value, end=" ")
    print("\n")
